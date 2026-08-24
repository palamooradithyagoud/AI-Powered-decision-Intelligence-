import openpyxl
import json
import os

wb = openpyxl.load_workbook('EMPLOYEE_ID.xlsx')
sheet = wb.active

avatar_colors = [
    "bg-indigo-600", "bg-purple-600", "bg-emerald-600", "bg-blue-600",
    "bg-teal-600", "bg-cyan-600", "bg-sky-600", "bg-pink-600",
    "bg-rose-600", "bg-amber-600", "bg-orange-600", "bg-violet-600"
]

employees_dict = {}

for i in range(2, sheet.max_row + 1):
    row = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)]
    if row[0] is not None:
        num = int(row[0])
        emp_id = str(row[1]).strip().lower()
        name = str(row[2]).strip()
        designation = str(row[3]).strip()
        skills = [s.strip() for s in str(row[4]).split(',') if s.strip()]
        exp_val = row[5]
        exp_years = float(exp_val) if isinstance(exp_val, (int, float)) else 3.0
        exp_str = f"{int(exp_years)} Years" if exp_years == int(exp_years) else f"{exp_years} Years"
        
        workload_raw = row[6]
        if isinstance(workload_raw, (int, float)):
            workload = int(round(float(workload_raw) * 100))
        else:
            workload = 50
            
        avail_raw = str(row[7]).strip()
        avail_lower = avail_raw.lower()
        if 'avail' in avail_lower and 'not' not in avail_lower:
            band = max(0, 100 - workload)
            availability = f"Available ({band}% bandwidth)"
            avail_status = "Available"
        elif 'partial' in avail_lower:
            band = max(0, 100 - workload)
            availability = f"Partial ({band}% bandwidth)"
            avail_status = "Partial"
        else:
            availability = "Busy (Fully Allocated)"
            avail_status = "Not Available"
            
        prev_proj_raw = row[8]
        if prev_proj_raw:
            prev_projects = [p.strip() for p in str(prev_proj_raw).split(';') if p.strip()]
        else:
            prev_projects = []
            
        color = avatar_colors[(num - 1) % len(avatar_colors)]
        email = f"{emp_id}@company.ai"

        # Determine user role
        desig_lower = designation.lower()
        if "project manager" in desig_lower or num == 1:
            role = "manager"
        elif "product manager" in desig_lower or "software architect" in desig_lower:
            role = "project_lead"
        else:
            role = "employee"

        emp_record = {
            "id": emp_id,
            "serial_no": num,
            "name": name,
            "email": email,
            "designation": designation,
            "role": role,
            "skills": skills,
            "experience": exp_str,
            "experience_years": exp_years,
            "workload": workload,
            "availability_status": avail_status,
            "availability": availability,
            "prev_projects": prev_projects,
            "avatar_color": color,
            "password": emp_id
        }
        employees_dict[num] = emp_record

json_data_str = json.dumps(employees_dict, indent=4)

file_content = f'''"""
Employee Dataset - Loaded from EMPLOYEE_ID.xlsx
Contains all 40 employees with full roles, skills, workload, and login credentials.
"""
from typing import Dict, List, Any, Optional

# Raw mapping with string and integer key lookups
_RAW_DATA: Dict[str, Dict[str, Any]] = {json_data_str}

EMPLOYEES_DATA: Dict[int, Dict[str, Any]] = {{
    int(k): v for k, v in _RAW_DATA.items()
}}

def get_all_employees() -> List[Dict[str, Any]]:
    """Return list of all 40 employee records sorted by serial number."""
    return sorted(list(EMPLOYEES_DATA.values()), key=lambda x: x["serial_no"])

def get_employee_by_num(num: int) -> Optional[Dict[str, Any]]:
    """Retrieve an employee profile by their serial number (1-40)."""
    try:
        n = int(num)
        return EMPLOYEES_DATA.get(n)
    except (ValueError, TypeError):
        return None

def get_employee_by_id(emp_id: str) -> Optional[Dict[str, Any]]:
    """Retrieve an employee profile by employee ID (e.g. emp_01, emp_02, emp_1)."""
    if not emp_id:
        return None
    clean = str(emp_id).strip().lower()
    if clean.isdigit():
        return get_employee_by_num(int(clean))
    
    # Try exact match
    for emp in EMPLOYEES_DATA.values():
        if emp["id"].lower() == clean:
            return emp
            
    # Try normalized emp_XX match (e.g. emp_1 -> emp_01)
    if clean.startswith("emp_"):
        suffix = clean.split("_")[1]
        if suffix.isdigit():
            return get_employee_by_num(int(suffix))
            
    return None

def get_employee_by_email_or_name(query: str) -> Optional[Dict[str, Any]]:
    """Find employee by email, ID, or full name (case-insensitive)."""
    if not query:
        return None
    q = str(query).strip().lower()
    
    # Direct ID check
    by_id = get_employee_by_id(q)
    if by_id:
        return by_id
        
    # Email or name check
    for emp in EMPLOYEES_DATA.values():
        if emp["email"].lower() == q or emp["name"].lower() == q:
            return emp
        if q in emp["email"].lower() or q in emp["name"].lower():
            return emp
            
    return None

def authenticate_employee(identifier: str, password: str) -> Optional[Dict[str, Any]]:
    """
    Authenticate employee credentials.
    Identifier can be employee ID (emp_01), email (emp_01@company.ai), or name.
    Password can be their employee ID (emp_01), or standard default passwords.
    """
    if not identifier or not password:
        return None
        
    emp = get_employee_by_email_or_name(identifier)
    if not emp:
        return None
        
    pwd_clean = str(password).strip().lower()
    valid_passwords = {{
        emp["id"].lower(),
        f"emp_{{emp['serial_no']}}",
        f"emp_{{emp['serial_no']:02d}}",
        "password123",
        "password",
        "admin",
        "kuiper123"
    }}
    
    if pwd_clean in valid_passwords:
        return emp
        
    return None
'''

os.makedirs("backend/db", exist_ok=True)
with open("backend/db/employees_data.py", "w", encoding="utf-8") as f:
    f.write(file_content)

print(f"Successfully generated backend/db/employees_data.py with {len(employees_dict)} employees!")

